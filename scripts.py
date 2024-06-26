from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from docx import Document

import shutil


#копируем базовую таблицу и вставляем по указанному адресу
def script_1(src: str, dst: str) -> None:
    shutil.copy2(src=src, dst=dst)


def script_2(gen_excel_table_path: str, key_word: str, output_excel_table_path: str, week_num=1) -> None:
    #открытие таблицы с расписанием
    woorkbook_input = load_workbook(filename=gen_excel_table_path)
    ws1 = woorkbook_input.active
    #открытие таблицы, в которую будет записан результат
    woorkbook_base = load_workbook(filename=output_excel_table_path)
    ws2 = woorkbook_base.active
    
    #перебираем все номера строк с таблице
    for row in range(1, ws1.max_row+1):
        #перебираем все номера столбцов в таблице
        for col in range(1, ws1.max_column+1):
            #получение ячейки: row-№строки, col-№столбца
            cell = ws1._get_cell(row, col)
            #получение значения ячейки
            value = str(cell.value)
            value_ed = del_all_spaces(value)

            #проверка наличия ключевого слова в ячейке (фамилия преподавателя)
            if key_word in value.split():
                #время пары
                time = del_all_spaces(ws1._get_cell(row=cell.row, column=2).value)
                #название группы
                group = get_group(del_spaces(ws1._get_cell(row=cell.row-1-offset[time], column=cell.column).value))
                #день недели
                day = del_spaces(ws1._get_cell(row=cell.row-offset[time], column=1).value)

                #определяет, есть ли разделение пар по неделям
                if '1н' in value_ed or '2н' in value_ed:
                    values = value.split('2 н')
                    if len(values) == 1:
                        values = value.split('2н')

                    #если пара с номером недели в расписании не соответсвует номеру нужной недели, изменения не вносятся, ячейка пропускается
                    if (key_word not in values[0] and week_num == 1) or (key_word not in values[1] and week_num == 2):
                        continue
                
                #base_days_1[day]: из словаря по ключу "day" получаем букву столбца таблицы
                #base_times[time]: из словаря по ключу "time" получаем номер строки
                #ws2[f'{base_days[day]}{base_times[time]}'] - ячейка в таблице вывода
                if ws2[f'{base_days_1[day]}{base_times[time]}'].value == None:
                    #если ячейка пустая, устанавливаем значение: группа
                    ws2[f'{base_days_1[day]}{base_times[time]}'] = f'{group}\n'
                    ws2[f'{base_days_2[day]}{base_times[time]}'] = f'{group}\n'
                else:
                    #если в ячейке уже есть группа, добавляем еще одну
                    ws2[f'{base_days_1[day]}{base_times[time]}'] = ws2[f'{base_days_1[day]}{base_times[time]}'].value+f'{group}\n'
                    ws2[f'{base_days_2[day]}{base_times[time]}'] = ws2[f'{base_days_2[day]}{base_times[time]}'].value+f'{group}\n'
                #сохраняем таблицу
                woorkbook_base.save(filename=output_excel_table_path)
    #закрываем соединение с таблицами
    woorkbook_input.close()
    woorkbook_base.close()


#удаление лишних пробелов из строки
def del_spaces(string: str) -> str:
    if string != None:
        return ' '.join([str(i) for i in string.split()])

#удаление всех пробелов из строки
def del_all_spaces(string: str) -> str:
    if string != None:
        return string.replace(' ', '')

#удаление лишних символов из названия группы (из общей таблицы с расписанием)
def get_group(group_name: str) -> str:
    group = group_name.split('(')[-1]
    return group[:-1]


def script_3(document_path: str, key_word: str, output_excel_table_path: str) -> None:
    #открытие конечной таблицы
    wb = load_workbook(output_excel_table_path)
    ws = wb.active
    #создание экземпляра документа
    doc = Document(document_path)
    #получение списка всех таблиц из документа
    tables = doc.tables
    #заводим счетчик, по значению которого можно будет отпределить день недели (букву столбца в талице вывода)
    #используя значение счетчика, как ключ в словаре "doc_days"
    counter = 0

    #перебираем все таблицы в документе
    for table in tables:
        counter += 1
        #перебираем все строки в таблице
        for row in table.rows:
            cells_list = row.cells

            if key_word in cells_list[3].text.split() and cells_list[3].text != cells_list[5].text:
                #получаем группу из word документа для которой есть замена 
                group_curr = get_correct_group(cells_list[1].text)
                #получаем список групп по расписанию из конечной таблицы
                group_list = ws[f'{doc_days[counter]}{int(cells_list[0].text)+1}'].value

                if group_list != None:
                    group_list = ws[f'{doc_days[counter]}{int(cells_list[0].text)+1}'].value.split('\n')
                else:
                    group_list = ''

                #убираем группу из расписания
                if group_curr in group_list:
                    #задает стиль ячейки (красный)
                    font = Font(color='52181b', bold=True)
                    fill = PatternFill(patternType='solid', fgColor='ff707a')
                    ws[f'{doc_days[counter]}{int(cells_list[0].text)+1}'].font = font
                    ws[f'{doc_days[counter]}{int(cells_list[0].text)+1}'].fill = fill
                    group_list.remove(group_curr)
                    text = '\n'.join(group for group in group_list)
                    ws[f'{doc_days[counter]}{int(cells_list[0].text)+1}'] = text

            elif key_word in cells_list[5].text.split() and cells_list[3].text != cells_list[5].text:
                #добавляет пару в рассписание
                #группа, которую надо добавить
                group_curr = get_correct_group(cells_list[1].text)
                if '-' not in cells_list[0].text:
                    #группы, которые уже есть в ячейке
                    text_old = ws[f'{doc_days[counter]}{int(cells_list[0].text)+1}'].value
                else:
                    continue

                if text_old != None:
                    ws[f'{doc_days[counter]}{int(cells_list[0].text)+1}'] = text_old+f'{group_curr} \n'
                else:
                    ws[f'{doc_days[counter]}{int(cells_list[0].text)+1}'] = f'{group_curr} \n'
                
                #задает стиль ячейки (зеленый)
                font = Font(color='1b4715', bold=True)
                fill = PatternFill(patternType='solid', fgColor='86d980')
                ws[f'{doc_days[counter]}{int(cells_list[0].text)+1}'].font = font
                ws[f'{doc_days[counter]}{int(cells_list[0].text)+1}'].fill = fill

    #сохранение изменений в таблице
    wb.save(output_excel_table_path)
    wb.close()

def get_correct_group(group_name: str) -> str:
    group_name = group_name.split()
    return f'{group_name[0]} - {group_name[-1]}'

def script_4(output_excel_table_path: str) -> None:
    #открытие таблицы вывода
    wb = load_workbook(output_excel_table_path)
    ws = wb.active

    for col in range(3, 12, 2):
        for row in range(2, 8):
            #значение текущей ячейки
            cell_text = ws._get_cell(row=row, column=col).value
            #значение предыдущей ячейки
            prev_cell_text = ws._get_cell(row=row, column=col-1).value

            if cell_text != None and prev_cell_text != None:
                #список групп в текущей ячейке
                cell_list = cell_text.split('\n')
                #список групп в предыдущей ячейке
                prev_cell_list = prev_cell_text.split('\n')

                for group in prev_cell_list:
                    if group not in cell_list:
                        for group in cell_list:
                            if group not in prev_cell_list:
                                #задает стиль ячейки (синий)
                                font = Font(color='000000', bold=True)
                                fill = PatternFill(patternType='solid', fgColor='427bf5')
                                ws._get_cell(row=row, column=col).font = font
                                ws._get_cell(row=row, column=col).fill = fill
                                wb.save(output_excel_table_path)

    #закрываем соединение с таблицей 
    wb.close()


offset = {'08.45-10.15': 0, '10.30-12.00': 1, '12.40-14.10': 2, '14.20-15.50': 3, '16.00-17.30': 4, '17.40-19.10': 5}
#используется для определения столбца по названию дня
base_days_1 = {'понедельник': 'B', 'вторник': 'D', 'среда': 'F', 'четверг': 'H', 'пятница': 'J'}
base_days_2 = {'понедельник': 'C', 'вторник': 'E', 'среда': 'G', 'четверг': 'I', 'пятница': 'K'}
#используется для определения номера строки в конечной таблице по времени пары
base_times = {'08.45-10.15': 2, '10.30-12.00': 3, '12.40-14.10': 4, '14.20-15.50': 5, '16.00-17.30': 6, '17.40-19.10': 7}
#писпользуется для определения столбца в конечной таблице по номеру таблицы в ворд-документе
doc_days = {1: 'C', 2: 'E', 3: 'G', 4: 'I', 5: 'K'}